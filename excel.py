import openpyxl as xl


inv_file = xl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}


for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inv_per_supplier = product_list.cell(product_row, 2).value

    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        product_per_supplier[supplier_name] = 1

    if supplier_name in total_value_per_supplier:
        current_num_inv = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_num_inv + inv_per_supplier
    else:
        total_value_per_supplier[inv_per_supplier] = 1


print(product_per_supplier)
print(inv_per_supplier)
